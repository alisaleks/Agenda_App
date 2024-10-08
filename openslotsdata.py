import pandas as pd
import pytz
from datetime import datetime, timedelta
import calendar 
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import time
import re

# Function to handle out-of-bound datetime values
def handle_out_of_bound_dates(date_str):
    try:
        return pd.to_datetime(date_str)
    except (pd.errors.OutOfBoundsDatetime, OverflowError):
        return pd.Timestamp.max

# Function to check if a resource is active
def is_active(row, start_date, end_date):
    if pd.isnull(row['EffectiveEndDate']) or pd.isnull(row['EffectiveStartDate']):
        return False
    return not (row['EffectiveEndDate'] < start_date or row['EffectiveStartDate'] > end_date)

def load_excel(file_path, usecols=None, **kwargs):
    # Load specific columns if usecols is provided to reduce memory usage
    return pd.read_excel(file_path, usecols=usecols, **kwargs)

def load_csv(file_path, usecols=None, **kwargs):
    # Load specific columns if usecols is provided to reduce memory usage
    return pd.read_csv(file_path, usecols=usecols, **kwargs)

shifts_columns_to_string = {
'Shift[ShiftNumber]': str,
'Shift[Label]': str,
'Service Resource[Name]': str,
'Shop[GT_ShopCode__c]': str,
'Service Resource[GT_Role__c]': str,
'Shift[StartTime]': str,
'Shift[EndTime]': str,
'Shift[ServiceResourceId]': str,
'Shop[GT_CountryCode__c]': str,
'Shop[Country]': str,
'Shop[Name]': str,
'Shop[GT_AreaManagerCode__c]': str,
'Shift[LastModifiedDate]': str,
'Service Resource[GT_PersonalNumber__c]': str,
'Shop[GT_StoreType__c]': str
}
resources_columns_to_string = {
'Shop[GT_CountryCode__c]': str,
'Shop[Country]': str,
'Service Territory Member[ServiceTerritoryId]': str,
'Shop[GT_ShopCode__c]': str,
'Service Resource[Name]': str,
'Service Territory Member[ServiceResourceId]': str,
'Service Resource[Name].1': str,
'Service Territory Member[EffectiveStartDate]': str,
'Service Territory Member[EffectiveEndDate]': str,
'Service Resource[GT_Role__c]': str,
'Service Resource[IsActive]' : str,
'Service Resource[GT_PersonalNumber__c]': str
}
appointments_columns_to_string = {
'Service Appointment[AppointmentNumber]': str,
'Service Appointment[ServiceTerritoryId]': str,
'Service Appointment[Business_Shop__c]': str,
'Service Appointment[GT_ShopCode__c]': str,
'Shop[GT_CountryCode__c]': str,
'Service Appointment[GT_Cluster__c]': str,
'Service Appointment[GT_Macrocategory__c]': str,
'Service Appointment[GT_AccountNameConcatenated__c]': str,
'Shop[GT_AreaCode__c]': str,
'Shop[GT_StoreType__c]': str,
'Shop[GT_AreaManagerCode__c]': str,
'Service Appointment[SchedStartTime]': str,
'Service Appointment[SchedEndTime]': str,
'Service Resource[GT_Role__c]': str,
'Service Appointment[GT_ServiceResource__c]': str,
'Service Resource[Name]': str,
'Service Appointment[Status]': str,
'Service Appointment[LastModifiedDate]': str
}
absences_columns_to_string = {
'Service Resource[GT_PersonalNumber__c]': str,
'User[GT_StoreCode__c]': str,
'Resource Absence[AbsenceNumber]': str,
'Service Resource[Name]': str,
'Service Resource[Id]':str,
'Resource Absence[Type]': str
}

# Load datasets with only the necessary columns specified
sfshifts = load_excel(
    'C:/Users/aaleksan/OneDrive - Amplifon S.p.A/Documentos/python_alisa/saturation/Saturation/Satapp/agenda_app/SFshifts_query.xlsx', 
    dtype=shifts_columns_to_string,
    usecols=[
        'Shift[ShiftNumber]', 'Shift[Label]', 'Service Resource[Name]', 'Shop[GT_ShopCode__c]', 
        'Service Resource[GT_Role__c]', 'Shift[StartTime]', 'Shift[EndTime]', 
        'Shift[ServiceResourceId]', 'Shop[GT_CountryCode__c]', 'Shop[Country]', 
        'Shop[Name]', 'Shop[GT_AreaManagerCode__c]', 'Shift[LastModifiedDate]', 
        'Service Resource[GT_PersonalNumber__c]', 'Shop[GT_StoreType__c]', 'Shop[GT_AreaCode__c]'
    ]
)

resources = load_csv(
    'C:/Users/aaleksan/OneDrive - Amplifon S.p.A/Documentos/python_alisa/saturation/Saturation/Satapp/agenda_app/resource_query.csv',  
    dtype=resources_columns_to_string,
    usecols=[
        'Shop[GT_CountryCode__c]', 'Service Territory Member[EffectiveEndDate]', 
        'Service Territory Member[EffectiveStartDate]', 'Shop[Country]', 
        'Service Territory Member[ServiceTerritoryId]', 'Shop[GT_ShopCode__c]', 
        'Service Territory Member[ServiceResourceId]', 'Service Resource[GT_PersonalNumber__c]', 'Service Resource[IsActive]',
        'Service Resource[GT_Role__c]','Service Resource[Name]'
    ], 
)

appointments = load_excel(
    'C:/Users/aaleksan/OneDrive - Amplifon S.p.A/Documentos/python_alisa/saturation/Saturation/Satapp/agenda_app/Appointments_aug_oct.xlsx', 
    dtype=appointments_columns_to_string,
    usecols=[
        'Service Appointment[AppointmentNumber]', 'Service Appointment[ServiceTerritoryId]', 
        'Service Appointment[Business_Shop__c]', 'Service Appointment[GT_ShopCode__c]', 
        'Shop[GT_CountryCode__c]', 'Service Appointment[GT_Cluster__c]', 
        'Service Appointment[GT_Macrocategory__c]', 'Service Appointment[GT_AccountNameConcatenated__c]', 
        'Shop[GT_AreaCode__c]', 'Shop[GT_StoreType__c]', 'Shop[GT_AreaManagerCode__c]', 
        'Service Appointment[SchedStartTime]', 'Service Appointment[SchedEndTime]', 
        'Service Resource[GT_Role__c]', 'Service Appointment[GT_ServiceResource__c]', 
        'Service Resource[Name]', 'Service Appointment[Status]', 'Service Appointment[LastModifiedDate]'
    ]
)

absences = load_csv(
    'C:/Users/aaleksan/OneDrive - Amplifon S.p.A/Documentos/python_alisa/saturation/Saturation/Satapp/agenda_app/absences.csv',
    dtype=absences_columns_to_string,
    usecols=[
        'Resource Absence[AbsenceNumber]', 'Resource Absence[Start]', 'Resource Absence[End]', 'Service Resource[Name]', 
        'Service Resource[GT_PersonalNumber__c]', 'User[GT_StoreCode__c]', 'Service Resource[Id]','Resource Absence[Type]'

    ]
)
# Rename columns to match
absences.rename(columns={
    'Resource Absence[Start]': 'Start',
    'Resource Absence[End]': 'End',
    'Resource Absence[AbsenceNumber]':'AbsenceNumber',
    'Service Resource[Name]': 'Resource.Name',
    'Service Resource[GT_PersonalNumber__c]': 'Resource.GT_PersonalNumber__c', 
    'User[GT_StoreCode__c]': 'Resource.RelatedRecord.GT_StoreCode__c',
    'Resource Absence[Type]': 'Type'
}, inplace=True)

# Load regionmapping data
region_mapping_path = 'C:/Users/aaleksan/OneDrive - Amplifon S.p.A/Documentos/python_alisa/saturation/Saturation/Satapp/agenda_app/regionmapping.xlsx'
region_mapping = load_excel(region_mapping_path)
# Filter the original region_mapping DataFrame
region_mapping = region_mapping[region_mapping['SYM'] != 'N']

sfshifts['StartTime'] = pd.to_datetime(sfshifts['Shift[StartTime]'], errors='coerce')
sfshifts['EndTime'] = pd.to_datetime(sfshifts['Shift[EndTime]'], errors='coerce')
def get_first_iso_week_start_date_current_month():
    # Get the current date
    today = datetime.today()
    
    # Get the first day of the current month with the time set to 00:00:00
    first_day_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Check if the first day of the month is a Sunday
    if first_day_of_month.weekday() == 6:  # Sunday is represented by 6 in weekday()
        # If Sunday, move to the next Monday
        first_day_of_month += timedelta(days=1)
    
    # Get the ISO calendar week and weekday of the first day of the month
    iso_year, iso_week, iso_weekday = first_day_of_month.isocalendar()
    
    # Calculate the difference to get back to the Monday of that ISO week
    # ISO weeks start on Monday (iso_weekday = 1), so subtract the days to go back to Monday
    start_date = first_day_of_month - timedelta(days=iso_weekday - 1)
    
    # Ensure the time part is set to 00:00:00
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    
    return start_date

def get_last_iso_week_end_date_current_month():
    # Get the current date
    today = datetime.today()
    
    # Get the last day of the current month with the time set to 00:00:00
    last_day_of_month = today.replace(day=calendar.monthrange(today.year, today.month)[1], hour=0, minute=0, second=0, microsecond=0)
    
    # Get the ISO calendar week and weekday of the last day of the month
    iso_year, iso_week, iso_weekday = last_day_of_month.isocalendar()
    
    # Calculate the difference to get to Sunday of that ISO week
    # ISO weeks end on Sunday (iso_weekday = 7), so add the days to go to Sunday
    end_date = last_day_of_month + timedelta(days=(7 - iso_weekday))
    
    # Ensure the time part is set to 00:00:00
    end_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
    
    return end_date

# Example usage: dynamically calculate start and end dates for the current month
start_date = get_first_iso_week_start_date_current_month()
end_date = get_last_iso_week_end_date_current_month()

print(f"The start date of the 1st ISO week of the current month (excluding Sunday start) is: {start_date}")
print(f"The end date of the last ISO week of the current month is: {end_date}")

start_iso_year, start_iso_week, _ = start_date.isocalendar()
end_iso_year, end_iso_week, _ = end_date.isocalendar()

shifts_filtered = sfshifts[(sfshifts['StartTime'] >= start_date) & (sfshifts['EndTime'] <= end_date)].copy()
# Rename columns to match
shifts_filtered.rename(columns={
    'Shop[GT_ShopCode__c]': 'GT_ShopCode__c',
    'Service Resource[Name]': 'GT_ServiceResource__r.Name'
}, inplace=True)
print(shifts_filtered[shifts_filtered['GT_ShopCode__c'] == '81C'])

resources.rename(columns={
    'Shop[GT_ShopCode__c]': 'GT_ShopCode__c'
}, inplace=True)
# Convert specific columns to datetime
shifts_filtered['LastModifiedDate'] = pd.to_datetime(shifts_filtered['Shift[LastModifiedDate]'], errors='coerce')
appointments['ApptStartTime'] = pd.to_datetime(appointments['Service Appointment[SchedStartTime]'], errors='coerce').dt.tz_localize(None)
appointments['ApptEndTime'] = pd.to_datetime(appointments['Service Appointment[SchedEndTime]'], errors='coerce').dt.tz_localize(None)
appointments['ApptsLastModifiedDate'] = pd.to_datetime(appointments['Service Appointment[LastModifiedDate]'], errors='coerce').dt.tz_localize(None)
# Drop original datetime columns
shifts_filtered.drop(columns=['Shift[StartTime]', 'Shift[EndTime]', 'Shift[LastModifiedDate]'], inplace=True)
appointments.drop(columns=['Service Appointment[SchedStartTime]', 'Service Appointment[SchedEndTime]', 'Service Appointment[LastModifiedDate]'], inplace=True)

shifts_filtered['PersonalNumberKey'] = shifts_filtered['GT_ShopCode__c'] + '_' + shifts_filtered['Service Resource[GT_PersonalNumber__c]']
a=shifts_filtered[ shifts_filtered['GT_ShopCode__c'] == '033']
a['Service Resource[GT_PersonalNumber__c]'].unique()
# Convert to datetime with out-of-bound handling for specific columns
resources['EffectiveEndDate'] = resources['Service Territory Member[EffectiveEndDate]'].apply(handle_out_of_bound_dates)
resources['EffectiveStartDate'] = resources['Service Territory Member[EffectiveStartDate]'].apply(handle_out_of_bound_dates)

# Add a date column
shifts_filtered['date'] = shifts_filtered['StartTime'].dt.strftime('%d/%m/%Y')
shifts_filtered['ShiftDate'] = shifts_filtered['StartTime'].dt.date
# Directly extract ISO week and year from StartTime
shifts_filtered['iso_week'] = shifts_filtered['StartTime'].dt.isocalendar().week
shifts_filtered['iso_year'] = shifts_filtered['StartTime'].dt.isocalendar().year

shifts_filtered['StartDateHour'] = shifts_filtered['StartTime'].dt.strftime('%Y-%m-%d %H:00:00')
shifts_filtered['Key'] = shifts_filtered['GT_ShopCode__c'] + '_' + shifts_filtered['GT_ServiceResource__r.Name'] + '_' + shifts_filtered['StartDateHour']
#duplicate treatment
duplicates = shifts_filtered[shifts_filtered.duplicated(subset=['Key'], keep=False)]
shifts_filtered = shifts_filtered.sort_values(by=['Key', 'LastModifiedDate'], ascending=[True, False])
shifts_filtered = shifts_filtered.drop_duplicates(subset=['Key'], keep='first')
shifts_filtered['ShiftDurationHours'] = (shifts_filtered['EndTime'] - shifts_filtered['StartTime']).dt.total_seconds() / 3600
shifts_filtered[(shifts_filtered['GT_ShopCode__c'] == '022') & (shifts_filtered['date'] == '02/10/2024')].head(100)
shifts_filtered['ShopResourceKey'] = shifts_filtered['GT_ShopCode__c'] + shifts_filtered['Shift[ServiceResourceId]']
resources['ShopResourceKey'] = resources['GT_ShopCode__c'] + resources['Service Territory Member[ServiceResourceId]']

check = shifts_filtered[(shifts_filtered['PersonalNumberKey'] == '022_40211') & (pd.to_datetime(shifts_filtered['ShiftDate'], errors='coerce') == '2024-10-02')]
check
# Step 1: Convert 'EffectiveStartDate' to datetime
resources['EffectiveStartDate'] = pd.to_datetime(resources['Service Territory Member[EffectiveStartDate]'], errors='coerce')

# Step 2: Sort resources by 'ShopResourceKey' and 'EffectiveStartDate' (latest first)
resources_sorted = resources.sort_values(by=['ShopResourceKey', 'EffectiveStartDate'], ascending=[True, False])

# Step 3: Drop duplicates in 'resources', keeping the latest 'EffectiveStartDate' for each 'ShopResourceKey'
resources_unique = resources_sorted.drop_duplicates(subset=['ShopResourceKey'], keep='first')

# Step 4: Add 'Active' status based on date range
resources_unique['Active'] = resources_unique.apply(is_active, axis=1, args=(start_date, end_date))
# Step 5: Merge 'shifts_filtered' with 'resources_unique' (add both 'Service Resource[IsActive]' and 'Active')
shifts_filtered = shifts_filtered.merge(
    resources_unique[['ShopResourceKey', 'Service Resource[IsActive]', 'Active']],
    on='ShopResourceKey',
    how='left'
)

# Step 6: Create a new key combining 'ShopResourceKey' and 'Shift[ShiftNumber]' to differentiate shifts
shifts_filtered['UniqueShiftKey'] = shifts_filtered['ShopResourceKey'] + '_' + shifts_filtered['Shift[ShiftNumber]']

# Step 7: Filter for only active resources
shifts_filtered = shifts_filtered[(shifts_filtered['Service Resource[IsActive]'] == 'True') & (shifts_filtered['Active'] == True)]

# Step 8: Remove duplicates based on the 'UniqueShiftKey'
shifts_filtered = shifts_filtered.drop_duplicates(subset=['UniqueShiftKey'])

# Step 9: Check the filtered data for 'PersonalNumberKey' and 'ShiftDate'
check = shifts_filtered[(shifts_filtered['PersonalNumberKey'] == '022_40211') & (pd.to_datetime(shifts_filtered['ShiftDate'], errors='coerce') == '2024-10-02')]

print(check)

shifts_filtered.columns

shifts_filtered['ShiftDurationHours'] = shifts_filtered['ShiftDurationHours'].fillna(0)

absences['Start'] = pd.to_datetime(absences['Start'], errors='coerce')
absences['End'] = pd.to_datetime(absences['End'], errors='coerce')
absences['PersonalNumberKey'] = absences['Resource.RelatedRecord.GT_StoreCode__c'] + '_' + absences['Resource.GT_PersonalNumber__c']

# Group absences by PersonalNumberKey and date to find total absence hours per day per resource
absences['AbsenceDate'] = absences['Start'].dt.date
# Modify the filtering logic to account for absences that overlap with the start_date and end_date
absences_filtered = absences[(absences['End'] >= start_date) & (absences['Start'] <= end_date)]

def expand_multiday_absences(row):
    start_date = row['Start'].normalize()
    end_date = row['End'].normalize()
    expanded_records = []
    current_date = start_date

    # Set the time threshold (20:00)
    evening_cutoff = pd.Timestamp(current_date).replace(hour=20, minute=0, second=0)

    while current_date <= end_date:
        if current_date == start_date and current_date == end_date:
            # Absence starts and ends on the same day
            if row['Start'] > evening_cutoff:
                hours = 0  # No absence counted if start is after 20:00
            else:
                hours = (row['End'] - row['Start']).total_seconds() / 3600
        elif current_date == start_date:
            # First day: Calculate hours from start time to midnight
            if row['Start'] > evening_cutoff:
                hours = 0  # No absence counted if start is after 20:00
            else:
                end_of_day = pd.Timestamp.combine(current_date + timedelta(days=1), pd.Timestamp.min.time())  # Midnight of the next day
                hours = (end_of_day - row['Start']).total_seconds() / 3600
        elif current_date == end_date:
            # Last day: Calculate hours from midnight to the end time
            start_of_day = pd.Timestamp.combine(current_date, pd.Timestamp.min.time())  # Midnight of the current day
            hours = (row['End'] - start_of_day).total_seconds() / 3600
        else:
            # Full day: Assign 24 hours for full days within the absence period, capped at 8 hours
            hours = 24

        # Cap hours to a maximum of 8 per day
        hours = min(8, hours)
        
        expanded_records.append({
            'PersonalNumberKey': row['PersonalNumberKey'],
            'AbsenceDate': current_date.date(),
            'AbsenceStartTime': row['Start'],
            'AbsenceEndTime': row['End'],
            'AbsenceNumber': row['AbsenceNumber'],
            'Resource.GT_PersonalNumber__c': row['Resource.GT_PersonalNumber__c'],
            'Resource.RelatedRecord.GT_StoreCode__c': row['Resource.RelatedRecord.GT_StoreCode__c'],
            'Resource.Name': row['Resource.Name'],
            'Service Resource[Id]': row['Service Resource[Id]'],
            'Type': row['Type']
        })
        current_date += timedelta(days=1)
    
    return expanded_records


# Filter absences to include any absence overlapping with the period
expanded_absences = absences_filtered.apply(expand_multiday_absences, axis=1)
expanded_absences = pd.DataFrame([record for sublist in expanded_absences for record in sublist])
expanded_absences.head()
# Group expanded absences by PersonalNumberKey and AbsenceDate
absences_grouped = expanded_absences.groupby(['PersonalNumberKey', 'AbsenceDate','AbsenceNumber']).agg({
    'Resource.GT_PersonalNumber__c': 'first', 
    'Resource.RelatedRecord.GT_StoreCode__c': 'first',  
    'Type': 'first',
    'Service Resource[Id]':'first',
    'Resource.Name': 'first',  
    'AbsenceStartTime': 'first',
    'AbsenceEndTime': 'last'
}).reset_index()
absences_grouped[(absences_grouped['Resource.GT_PersonalNumber__c'] == '33104')].head(25)
# Now perform the filtering with the correct date type
target_date = datetime.strptime('2024-10-03', '%Y-%m-%d').date()

filtered_absences = absences_grouped[
    (absences_grouped['PersonalNumberKey'] == '003_11126') & 
    (absences_grouped['AbsenceDate'] == target_date)
].head(60)
filtered_absences[['PersonalNumberKey', 'AbsenceDate','Resource.Name', 'AbsenceStartTime',
       'AbsenceEndTime']]
# Group shifts by PersonalNumberKey and ShiftDate to find total shift hours per day per resource
shifts_grouped = shifts_filtered.groupby(['PersonalNumberKey', 'ShiftDate']).agg({
    'ShiftDurationHours': 'sum',  # Sum of absence duration hours
    'Service Resource[GT_Role__c]' : 'first', 
    'Shift[Label]' : 'first',
    'GT_ServiceResource__r.Name' : 'first',
    'GT_ShopCode__c': 'first',
    'ShopResourceKey': 'first',  
    'StartDateHour': 'first',  
    'iso_year': 'first',  
    'iso_week': 'first',
    'date': 'first',
    'LastModifiedDate' : 'first',
    'GT_ShopCode__c': 'first',
    'Shop[Name]': 'first',
    'Shop[GT_CountryCode__c]': 'first',
    'Shop[Country]': 'first',
    'Shop[GT_AreaManagerCode__c]': 'first',
    'Shop[GT_AreaCode__c]' : 'first',
    'Shop[GT_StoreType__c]': 'first',
    'StartTime': 'first',
    'EndTime': 'last'
    
}).reset_index()
shifts_grouped.head()
check= shifts_grouped[shifts_grouped['PersonalNumberKey'] == '003_11126']
check[['PersonalNumberKey', 'ShiftDate', 'ShiftDurationHours','StartTime', 'EndTime']]
# Step 1: Initialize an empty list to store expanded shifts
shift_slots_5mins = []
# Step 2: Loop through each row in sfshifts_merged and generate 5-minute slots
for _, row in shifts_grouped.iterrows():
    shift_start = row['StartTime']
    shift_end = row['EndTime']
    
    # Generate 5-minute slots between shift_start and shift_end
    while shift_start < shift_end:
        shift_slots_5mins.append({
            'ShopResourceKey': row['ShopResourceKey'],
            'PersonalNumberKey': row['PersonalNumberKey'],
            'ShiftSlot': shift_start,  # The start time of the slot
            'ShiftDurationHours': row['ShiftDurationHours'],
            'ShiftLabel': row['Shift[Label]'],
            'date': row['date'],
            'iso_week': row['iso_week'],
            'iso_year': row['iso_year'],
            'GT_ServiceResource__r.Name': row['GT_ServiceResource__r.Name'], 
            'GT_ShopCode__c': row['GT_ShopCode__c'],
            'Shop[Name]' : row['Shop[Name]'],
            'StartTime': row['StartTime'], 
            'EndTime': row['EndTime'],
            'LastModifiedDate': row['LastModifiedDate'],
            'StartDateHour': row['StartDateHour']
        })
        shift_start += timedelta(minutes=5)

# Step 3: Convert the list to a DataFrame
expanded_shifts_df = pd.DataFrame(shift_slots_5mins)
expanded_shifts_df[expanded_shifts_df['PersonalNumberKey'] == '003_11126'].head()
# Assuming 'StartTime' and 'EndTime' columns are added during the slot expansion process.
filtered_shifts_df = expanded_shifts_df[['ShiftSlot', 'StartTime', 'EndTime', 'ShiftLabel']]
expanded_shifts_df.columns
filtered_shifts = expanded_shifts_df[
    (expanded_shifts_df['PersonalNumberKey'] == '003_11126') & 
    (expanded_shifts_df['ShiftSlot'].dt.date == target_date)
].head(60)
filtered_shifts[['ShiftSlot', 'StartTime', 'EndTime', 'ShiftLabel']]


absence_slots_5mins = []

# Step 2: Loop through each row in absences_grouped and generate 5-minute slots
for _, row in absences_grouped.iterrows():
    absence_start = pd.Timestamp.combine(row['AbsenceDate'], row['AbsenceStartTime'].time()) if pd.notnull(row['AbsenceStartTime']) else row['AbsenceDate']
    absence_end = pd.Timestamp.combine(row['AbsenceDate'], row['AbsenceEndTime'].time()) if pd.notnull(row['AbsenceEndTime']) else row['AbsenceDate'] + timedelta(hours=8)
    
    # Generate 5-minute slots between absence_start and absence_end
    while absence_start < absence_end:
        absence_slots_5mins.append({
            'PersonalNumberKey': row['PersonalNumberKey'],
            'AbsenceSlot': absence_start,  # The start time of the absence slot
            'AbsenceNumber': row['AbsenceNumber'],
            'Resource.GT_PersonalNumber__c': row['Resource.GT_PersonalNumber__c'], 
            'GT_ShopCode__c':row['Resource.RelatedRecord.GT_StoreCode__c'],
            'Type': row['Type'],
            'Resource.Name': row['Resource.Name'],
            'Service Resource[Id]': row['Service Resource[Id]']
        })
        absence_start += timedelta(minutes=5)

# Step 3: Convert the list to a DataFrame
expanded_absences_df = pd.DataFrame(absence_slots_5mins)
expanded_absences_df['PersonalidKey'] = expanded_absences_df['GT_ShopCode__c'] + expanded_absences_df['Service Resource[Id]']
# The goal here is to filter out any absence that does not fall within a shift slot
# Now perform the filtering with the correct date type
filtered_absences_exp = expanded_absences_df[
    (expanded_absences_df['PersonalNumberKey'] == '003_11126') & 
    (expanded_absences_df['AbsenceSlot'].dt.date == target_date)
].head(60)
filtered_absences_exp
# Merge absence slots with shift slots on the PersonalNumberKey and 5-minute time slot
absences_with_shifts = pd.merge_asof(
    expanded_absences_df.sort_values('AbsenceSlot'),
    expanded_shifts_df[['PersonalNumberKey', 'ShiftSlot']].sort_values('ShiftSlot'),
    left_on='AbsenceSlot',
    right_on='ShiftSlot',
    by='PersonalNumberKey',
    tolerance=pd.Timedelta('5min'),
    direction='nearest'
)

# Now perform the filtering with the correct date type
filtered_absences_exp = absences_with_shifts[
    (absences_with_shifts['PersonalNumberKey'] == '003_11126') & 
    (absences_with_shifts['AbsenceSlot'].dt.date == target_date)
].head(60)
filtered_absences_exp
# Step 2: Drop duplicates if needed (optional, depending on your data)
absences_with_shifts = absences_with_shifts.drop_duplicates()
target_date = datetime.strptime('2024-10-03', '%Y-%m-%d').date()

# Now perform the filtering with the correct date type
filtered_absences = absences_grouped[
    (absences_grouped['PersonalNumberKey'] == '003_11126') & 
    (absences_grouped['AbsenceDate'] == target_date)
].head(60)
filtered_absences
absences_with_shifts['AbsenceSlotDate'] = absences_with_shifts['AbsenceSlot'].dt.date
# Rename the AbsenceSlotDate to 'date' to align with other datasets
# Include 'PersonalNumberKey' in the grouping to ensure correct counting for each person
absences_with_shifts_unique = absences_with_shifts.drop_duplicates(subset=['PersonalNumberKey', 'AbsenceSlot'])

# Step 2: Group by 'GT_ShopCode__c', 'AbsenceSlotDate', and 'PersonalNumberKey' to count unique absence slots
absence_slots = absences_with_shifts_unique.groupby(
    ['GT_ShopCode__c', 'AbsenceSlotDate', 'PersonalNumberKey']
).agg({
    'AbsenceSlot': 'count',  # Count the number of unique absence slots
    'AbsenceNumber': 'first',
    'Resource.GT_PersonalNumber__c': 'first',
    'Type': 'first',
    'Resource.Name': 'first',
    'Service Resource[Id]': 'first',
    'PersonalidKey': 'first',
    'ShiftSlot': 'first',
}).reset_index()

# Rename 'AbsenceSlot' column to 'AbsenceSlots' for clarity
absence_slots.rename(columns={'AbsenceSlot': 'AbsenceSlots'}, inplace=True)


# Filter for a specific shop and date
target_shop = '994'
filtered_sfshifts_merged = absence_slots[
    (absence_slots['GT_ShopCode__c'] == target_shop) 
]

# Select only the specified columns
filtered_columns = ['AbsenceSlotDate', 'AbsenceSlots', 'PersonalNumberKey']
filtered_sfshifts_merged = filtered_sfshifts_merged[filtered_columns]

# Display the result
filtered_sfshifts_merged
shifts_grouped = shifts_grouped.groupby(['PersonalNumberKey', 'ShiftDate']).agg({
    'ShiftDurationHours': 'sum',  # Sum of absence duration hours
    'Service Resource[GT_Role__c]' : 'first', 
    'GT_ServiceResource__r.Name' : 'first',
    'GT_ShopCode__c': 'first',
    'Shift[Label]' : 'first',
    'ShopResourceKey': 'first',  
    'StartDateHour': 'first',  
    'iso_year': 'first',  
    'iso_week': 'first',
    'date': 'first',
    'LastModifiedDate' : 'first',
    'GT_ShopCode__c': 'first',
    'Shop[Name]': 'first',
    'Shop[GT_CountryCode__c]': 'first',
    'Shop[Country]': 'first',
    'Shop[GT_AreaManagerCode__c]': 'first',
    'Shop[GT_AreaCode__c]' : 'first',
    'Shop[GT_StoreType__c]': 'first',
    'StartTime': 'first',
    'EndTime': 'last'
    
}).reset_index()

# Merge expanded_absences with shifts data to calculate adjusted shift hours
sfshifts_merged = pd.merge(
    shifts_grouped,
    absence_slots,
    how='left',
    left_on=['PersonalNumberKey', 'ShiftDate'],
    right_on=['PersonalNumberKey','AbsenceSlotDate'],
    suffixes=('', '_absence')
)
target_shop = '994'

# Filter based on shop and date
filtered_sfshifts_merged = sfshifts_merged[
    (sfshifts_merged['GT_ShopCode__c'] == target_shop) ]
# Select only the specified columns
filtered_columns = [
    'AbsenceSlotDate', 'AbsenceSlots',
    'PersonalNumberKey', 'ShiftDate', 'ShiftDurationHours'
]

filtered_sfshifts_merged = filtered_sfshifts_merged[filtered_columns]

filtered_sfshifts_merged
sfshifts_merged['AbsenceSlots'] = sfshifts_merged['AbsenceSlots'].fillna(0)

sfshifts_merged['AbsenceDurationHours'] = sfshifts_merged['AbsenceSlots'] * 5 / 60

# Adjust AbsenceDurationHours if ShiftDurationHours is greater
sfshifts_merged['AbsenceDurationHours'] = sfshifts_merged.apply(
    lambda row: row['ShiftDurationHours'] if row['ShiftDurationHours'] < row['AbsenceDurationHours'] else row['AbsenceDurationHours'],
    axis=1
)

# Calculate the adjusted shift duration by subtracting the absence hours
sfshifts_merged['ShiftDurationHoursAdjusted'] = sfshifts_merged['ShiftDurationHours'] - sfshifts_merged['AbsenceDurationHours'].fillna(0)
# replacing negatives to 0
sfshifts_merged['ShiftDurationHoursAdjusted'] = sfshifts_merged['ShiftDurationHoursAdjusted'].apply(lambda x: max(x, 0))
# Recalculate ShiftDurationMinutes based on adjusted hours
sfshifts_merged['ShiftDurationMinutesAdjusted'] = sfshifts_merged['ShiftDurationHoursAdjusted'] * 60
sfshifts_merged['ShiftDate'] = pd.to_datetime(sfshifts_merged['ShiftDate'])

check = sfshifts_merged[(sfshifts_merged['GT_ShopCode__c'] == '88F') & (pd.to_datetime(sfshifts_merged['ShiftDate'], errors='coerce') == '2024-09-18')]
columns_to_display = ['ShiftDurationHours', 'AbsenceDurationHours',  'Shop[Name]', 'Resource.Name']
# Filter the DataFrame and select only the required columns
check_filtered = check[columns_to_display]

# Identify duplicates based on the specified subset of columns
duplicate_mask = appointments.duplicated(subset=[
    'Shop[GT_CountryCode__c]',
    'Service Appointment[GT_ShopCode__c]',
    'Service Resource[Name]',
    'Service Appointment[GT_AccountNameConcatenated__c]',
    'ApptStartTime',
    'ApptEndTime'
], keep=False)

# Filter the DataFrame to show only duplicates
duplicates = appointments[duplicate_mask]

# Sort by the specified subset of columns and 'ApptsLastModifiedDate'
appointments = appointments.sort_values(by=[
    'Shop[GT_CountryCode__c]',
    'Service Appointment[GT_ShopCode__c]',
    'Service Resource[Name]',
    'Service Appointment[GT_AccountNameConcatenated__c]',
    'ApptStartTime',
    'ApptEndTime',
    'ApptsLastModifiedDate'
], ascending=[True, True, True, True, True, True, False])

# Drop duplicates, keeping only the last modified
appointments = appointments.drop_duplicates(subset=[
    'Shop[GT_CountryCode__c]',
    'Service Appointment[GT_ShopCode__c]',
    'Service Resource[Name]',
    'Service Appointment[GT_AccountNameConcatenated__c]',
    'ApptStartTime',
    'ApptEndTime'
], keep='first')

# Filter appointments within August
appointments_filtered = appointments[(appointments['ApptStartTime'] >= start_date) & (appointments['ApptEndTime'] <= end_date)].copy()
# First, create a dictionary to map PersonalNumber to Employee Full Name in sfshifts_merged
resources.rename(columns={'Service Resource[GT_PersonalNumber__c]': 'PersonalNumber'}, inplace=True)
resources.columns
personal_number_to_name = {
    f"{resources['GT_ShopCode__c'][index]}_{personal_number}":resources['Service Resource[Name]'][index]
    for index, personal_number in enumerate(resources['PersonalNumber'])
}
#Deleting unused DataFrames to free up memory
del sfshifts, resources, appointments, absences

# Fill NA in categories with 'First Visit'
appointments_filtered['Service Appointment[GT_Macrocategory__c]'] = appointments_filtered['Service Appointment[GT_Macrocategory__c]'].fillna('First Visit')
appointments_filtered['Service Appointment[GT_Macrocategory__c]'] = appointments_filtered['Service Appointment[GT_Macrocategory__c]'].str.strip()
appointments_filtered['Service Appointment[GT_Macrocategory__c]'] = appointments_filtered['Service Appointment[GT_Macrocategory__c]'].replace({
    'Fitting': 'Pre-Sales',
    'Post-Sales': 'After-Sales'
})


# Step 1: Generate all 5-minute slots for all appointments across all categories
slots = []

# Step 1: Generate all 5-minute slots for all appointments across all categories
for _, row in appointments_filtered.iterrows():
    slot_start = row['ApptStartTime']
    slot_end = row['ApptEndTime']
    
    # Generate all 5-minute slots for this appointment
    while slot_start < slot_end:
        slots.append((row['Service Appointment[GT_ShopCode__c]'], row['Service Resource[Name]'], slot_start.strftime('%d/%m/%Y'), slot_start, row['ApptsLastModifiedDate'], row['ApptStartTime'], 
            row['ApptEndTime'], row['Service Appointment[GT_ServiceResource__c]']))
        slot_start += timedelta(minutes=5)

# Create a DataFrame from all the slots including the last modified date for duplicate removal
all_slots_df = pd.DataFrame(slots, columns=['GT_ShopCode__c', 'Service Resource[Name]', 'date', 'Slot', 'LastModifiedDate', 'ApptStartTime', 'ApptEndTime', 'Service Appointment[GT_ServiceResource__c]'])
# Ensure the 'date' column is of datetime type for proper merging
all_slots_df['date'] = pd.to_datetime(all_slots_df['date'], format='%d/%m/%Y')
all_slots_df['PersonalidKey'] = all_slots_df['GT_ShopCode__c'] + all_slots_df['Service Appointment[GT_ServiceResource__c]']
# Sort slots by 'GT_ShopCode__c', 'Service Resource[Name]', 'date', 'Slot', and 'LastModifiedDate' to prioritize keeping the earliest slot
all_slots_df = all_slots_df.sort_values(by=['GT_ShopCode__c', 'Service Resource[Name]', 'date', 'Slot', 'LastModifiedDate'], ascending=[True, True, True, True, False])

# Remove duplicate 5-minute slots, keeping only the earliest modified slot
all_slots_df = all_slots_df.drop_duplicates(subset=['GT_ShopCode__c', 'Service Resource[Name]', 'date', 'Slot'], keep='first')

# Step 1: Merge shift slots and appointment slots based on PersonalidKey, date, and Slot
# The goal here is to ensure only appointments that match an active shift are kept.
all_slots_df['Slot'] = pd.to_datetime(all_slots_df['Slot'])
expanded_shifts_df['ShiftSlot'] = pd.to_datetime(expanded_shifts_df['ShiftSlot'])

appointments_with_shifts = pd.merge(
    all_slots_df,
    expanded_shifts_df[['ShopResourceKey', 'ShiftSlot']],
    how='inner',
    left_on=['PersonalidKey', 'Slot'],
    right_on=['ShopResourceKey', 'ShiftSlot']
)

duplicates = appointments_with_shifts[appointments_with_shifts.duplicated(subset=['ShopResourceKey', 'ShiftSlot'], keep=False)]
duplicates.head(20)
# Drop duplicates based on 'ShopResourceKey' and 'ShiftSlot' columns
appointments_with_shifts = appointments_with_shifts.drop_duplicates(subset=['ShopResourceKey', 'ShiftSlot'], keep='first')

target_date = pd.to_datetime('2024-09-20')
personal_key = '1850Hn670000008WNrCAM'
appointments_example = all_slots_df[
    (all_slots_df['PersonalidKey'] == personal_key) & 
    (all_slots_df['date'] == target_date)
]
# Step 3: Merge with absence slots to check for overlaps
appointments_with_shifts_and_absences = pd.merge(
    appointments_with_shifts,
    absences_with_shifts[['PersonalidKey', 'AbsenceSlot']],
    how='left',
    left_on=['PersonalidKey', 'Slot'],
    right_on=['PersonalidKey', 'AbsenceSlot']
)
# Step 4: Prioritize appointments over absences
appointments_with_shifts_and_absences['IsAbsence'] = appointments_with_shifts_and_absences['AbsenceSlot'].notnull()
appointments_with_shifts_and_absences['IsAppointment'] = appointments_with_shifts_and_absences['Slot'].notnull()
# absences_with_shifts already contains all absences, so we'll use it directly
all_absence_slots = absences_with_shifts.copy()
appointments_with_shifts_and_absences
# Merge absence slots with appointment slots to identify overlaps
overlapping_absence_slots = pd.merge(
    absences_with_shifts[['PersonalidKey', 'AbsenceSlot']],
    appointments_with_shifts_and_absences[['PersonalidKey', 'Slot']],
    left_on=['PersonalidKey', 'AbsenceSlot'],
    right_on=['PersonalidKey', 'Slot'],
    how='inner'
)
overlapping_absence_slots.columns
appointments_with_shifts_and_absences[(appointments_with_shifts_and_absences['GT_ShopCode__c'] == '25018')].tail()
target_date = pd.to_datetime('2024-10-23')
absences_example = overlapping_absence_slots[
    (overlapping_absence_slots['PersonalidKey'] == '0100Hn67000000PNF9CAO') & 
    (overlapping_absence_slots['AbsenceSlot'].dt.date == target_date.date())
]
absences_example.tail(20)
# Only keep the necessary columns
overlapping_absence_slots = overlapping_absence_slots[['PersonalidKey', 'AbsenceSlot']]
# Add GT_ShopCode__c and AbsenceSlot date to the overlapping absence slots for grouping
overlapping_absence_slots = pd.merge(
    overlapping_absence_slots,
    absences_with_shifts[['PersonalidKey', 'GT_ShopCode__c', 'AbsenceSlot']],
    on=['PersonalidKey', 'AbsenceSlot'],
    how='left'
)

# Convert AbsenceSlot to date for grouping purposes
overlapping_absence_slots['AbsenceSlotDate'] = overlapping_absence_slots['AbsenceSlot'].dt.date
overlapping_duplicates = overlapping_absence_slots[overlapping_absence_slots.duplicated(subset=['PersonalidKey', 'AbsenceSlot'], keep=False)]
overlapping_duplicates.head(20)
# Drop duplicates based on 'ShopResourceKey' and 'ShiftSlot' columns
overlapping_absence_slots = overlapping_absence_slots.drop_duplicates(subset=['PersonalidKey', 'AbsenceSlot'], keep='first')

# Group by shop, service resource, and date to calculate the total overlapping absence slots
total_overlapping_absence_slots = overlapping_absence_slots.groupby(
    ['GT_ShopCode__c', 'AbsenceSlotDate']
).size().reset_index(name='TotalOverlappingAbsenceSlots')

# Rename the AbsenceSlotDate to 'date' to align with other datasets
total_overlapping_absence_slots.rename(columns={'AbsenceSlotDate': 'date'}, inplace=True)
total_overlapping_absence_slots[total_overlapping_absence_slots['GT_ShopCode__c'] == '010']
# Step 2: Count the unique slots, ensuring no double-counting for overlaps
net_booked_slots = appointments_with_shifts_and_absences.groupby(['GT_ShopCode__c', 'Service Resource[Name]', 'date', 'Slot']).size().reset_index(name='Count')

# Normalize counts to 1 for any slot count greater than 1
net_booked_slots['Count'] = net_booked_slots['Count'].apply(lambda x: 1 if x > 0 else 0)

# Calculate the total booked slots by shop, service resource, and date
total_booked_slots_by_date = net_booked_slots.groupby(['GT_ShopCode__c', 'Service Resource[Name]', 'date'])['Count'].sum().reset_index()
total_booked_slots_by_date.head()
total_booked_slots_by_date['date'] = pd.to_datetime(total_booked_slots_by_date['date'], errors='coerce')
grouped_df = total_booked_slots_by_date.groupby(['GT_ShopCode__c', 'date']).agg({'Count': 'sum'}).reset_index()
# Calculate the total number of 5-minute slots available per shop and date
total_overlapping_absence_slots['date'] = pd.to_datetime(total_overlapping_absence_slots['date'], errors='coerce')
shift_slots = sfshifts_merged.groupby(['GT_ShopCode__c', 'Shop[Name]', 'date'])[['ShiftDurationMinutesAdjusted', 'ShiftDurationHours','AbsenceDurationHours', 'ShiftDurationHoursAdjusted']].sum().reset_index()
shift_slots['date'] = pd.to_datetime(shift_slots['date'], format='%d/%m/%Y', errors='coerce')
shift_slots = pd.merge(shift_slots, total_overlapping_absence_slots, on=['GT_ShopCode__c', 'date'], how='left')
shift_slots['TotalOverlappingAbsenceSlots'] = shift_slots['TotalOverlappingAbsenceSlots'].fillna(0)
shift_slots['OverlapHours'] = (shift_slots['TotalOverlappingAbsenceSlots']* 5) / 60
shift_slots['TotalSlots_net'] = shift_slots['ShiftDurationMinutesAdjusted'] / 5
shift_slots['TotalSlots_net_gross'] = (shift_slots['ShiftDurationHours']*60) / 5
shift_slots['TotalHours'] = shift_slots['ShiftDurationHours'].fillna(0)
shift_slots['BlockedHours'] = shift_slots['AbsenceDurationHours'].fillna(0)-shift_slots['OverlapHours'].fillna(0)
shift_slots['BlockedHours'] = shift_slots['BlockedHours'].apply(lambda x: max(x, 0))
shift_slots['AvailableHours']= shift_slots['TotalHours'] - shift_slots['BlockedHours'] 
shift_slots['BlockedHoursPercentage'] = (shift_slots['BlockedHours'] / shift_slots['TotalHours']) * 100
# Convert the string '2024-09-24' to a datetime object for proper comparison
target_date = datetime.strptime('2024-09-24', '%Y-%m-%d')

# Correct the parentheses and comparison
filtered_shift_slots = shift_slots[
    (shift_slots['GT_ShopCode__c'] == '88F') & 
    (shift_slots['date'] == target_date)
]

filtered_shift_slots

# Step 1: Generate all dates within the specified range
date_range = pd.date_range(start=start_date, end=end_date, freq='B')  # weekdays only

# Step 2: Create a DataFrame for all combinations of shop codes and the date range
shops_dates = pd.MultiIndex.from_product(
    [region_mapping['CODE'].unique(), date_range],
    names=['GT_ShopCode__c', 'date']
).to_frame(index=False)

# Step 3: Include Region, Area, and Shop[Name] information in the shops_dates DataFrame
shops_dates = pd.merge(
    shops_dates,
    region_mapping[['CODE', 'REGION', 'AREA', 'DESCR']],  # Include Region, Area, and Shop[Name]
    left_on='GT_ShopCode__c',
    right_on='CODE',
    how='left'
)

# Rename columns to match the expected output
shops_dates.rename(columns={
    'REGION': 'REGION',
    'AREA': 'AREA',
    'DESCR': 'Shop[Name]'
}, inplace=True)

# Step 4: Drop unnecessary columns
shops_dates.drop(columns=['CODE'], inplace=True)

shift_slots = pd.merge(
    shops_dates,
    shift_slots,
    on=['GT_ShopCode__c', 'date'],
    how='left', 
    suffixes=('', '_drop')  # Use '_drop' as the suffix for the columns you want to drop
)
shift_slots = shift_slots.loc[:, ~shift_slots.columns.str.endswith('_drop')]
# Step 4: Fill missing values for any shops that had no shifts
shift_slots['ShiftDurationHours'] = shift_slots['ShiftDurationHours'].fillna(0)
shift_slots['ShiftDurationMinutesAdjusted'] = shift_slots['ShiftDurationMinutesAdjusted'].fillna(0)
shift_slots['ShiftDurationHoursAdjusted'] = shift_slots['ShiftDurationHoursAdjusted'].fillna(0)
shift_slots['AbsenceDurationHours'] = shift_slots['AbsenceDurationHours'].fillna(0)
shift_slots['TotalSlots_net'] = shift_slots['TotalSlots_net'].fillna(0)
shift_slots['TotalHours'] = shift_slots['TotalHours'].fillna(0)
shift_slots['BlockedHours'] = shift_slots['AbsenceDurationHours'].fillna(0)-shift_slots['OverlapHours'].fillna(0)
shift_slots['BlockedHoursPercentage'] = shift_slots['BlockedHoursPercentage'].fillna(0)
shift_slots['AvailableHours'] = shift_slots['AvailableHours'].fillna(0)

# Test shop 'C07' to check final output
a = shift_slots[shift_slots['GT_ShopCode__c'] == '81C']
a

# Step 5: Recalculate `TotalBookedSlots` based on the total booked slots by date
shift_slots = pd.merge(
    shift_slots, 
    grouped_df, 
    on=['GT_ShopCode__c', 'date'], 
    how='left'
)

# Fill missing values for TotalBookedSlots and perform necessary calculations
shift_slots['TotalBookedSlots'] = shift_slots['Count'].fillna(0)
shift_slots.drop(columns=['Count'], inplace=True)

# Step 6: Additional Calculations
shift_slots['BookedHours'] = (shift_slots['TotalBookedSlots'] * 5) / 60
shift_slots['SaturationPercentage'] = (shift_slots['BookedHours'] / shift_slots['TotalHours']) * 100
shift_slots['SaturationPercentage'] = shift_slots['SaturationPercentage'].clip(lower=0, upper=100)
shift_slots['OpenHours'] = shift_slots['TotalHours'] - shift_slots['BookedHours']-shift_slots['BlockedHours']
shift_slots['OpenHours'] = shift_slots['OpenHours'].apply(lambda x: max(x, 0))

# Add weekday name and ISO week
shift_slots['date'] = pd.to_datetime(shift_slots['date'], format='%d/%m/%Y')
shift_slots['day'] = shift_slots['date'].dt.day
shift_slots['weekday'] = shift_slots['date'].dt.day_name()
shift_slots['iso_week'] = shift_slots['date'].dt.isocalendar().week
shift_slots['month'] = shift_slots['date'].dt.strftime('%B')

# Remove Sundays if needed
shift_slots = shift_slots[shift_slots['weekday'] != 'Sunday']
shift_slots.columns
# Sort the shift_slots DataFrame by 'date'
shift_slots = shift_slots.sort_values(by='date')
shift_slots.rename(columns={
    'REGION': 'Region',
    'AREA': 'Area',
    'DESCR': 'Shop[Name]'
}, inplace=True)
# Save to Excel
current_date = datetime.now().strftime("%Y-%m-%d")
output_file_path = f'shiftslots_{current_date}.xlsx' 
shift_slots.to_excel(output_file_path, index=False, engine='openpyxl')

filtered_shift_slots = shift_slots[shift_slots['date_column'] == current_date]
output_file_path_today= f'hours_{current_date}.xlsx'
filtered_shift_slots.to_excel(output_file_path_today, index=False, engine='openpyxl')

# Now to convert it into a table format
wb = load_workbook(output_file_path_today)
ws = wb.active

# Define the range of the data to create a table
min_col = ws.min_column
max_col = ws.max_column
min_row = ws.min_row
max_row = ws.max_row

# Create a table reference for your data
table = Table(displayName="ShiftSlotsTable", ref=f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}")

# Add a default style to the table
style = TableStyleInfo(
    name="TableStyleMedium9", 
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True)
table.tableStyleInfo = style

# Add the table to the worksheet
ws.add_table(table)

# Save the workbook with the table
wb.save(output_file_path_today)

#TAB4
sfshifts_merged.head()
# Step 3: Include Region, Area, and Shop[Name] information in the shops_dates DataFrame
sfshifts_merged = pd.merge(
    sfshifts_merged,
    region_mapping[['CODE', 'REGION', 'AREA', 'DESCR']],  
    left_on='GT_ShopCode__c',
    right_on='CODE',
    how='left'
)

missing_shop_codes = sfshifts_merged[sfshifts_merged['GT_ShopCode__c'].isna()]
print(missing_shop_codes[['GT_ShopCode__c', 'Shop[Name]', 'REGION', 'AREA', 'CODE', 'ShiftDurationHoursAdjusted']])

# Rename columns to match the expected output
sfshifts_merged.rename(columns={
    'REGION': 'Region',
    'AREA': 'Area',
    'DESCR': 'Shop[Name]'
}, inplace=True)

# Step 4: Drop unnecessary columns
sfshifts_merged.drop(columns=['CODE'], inplace=True)
sfshifts_merged.fillna(0, inplace=True)

sfshifts_merged['weekday'] = sfshifts_merged['ShiftDate'].dt.day_name()
sfshifts_merged.columns

# Save to Excel
output_file_path2 = 'hcpshiftslots.xlsx'
sfshifts_merged.to_excel(output_file_path2, index=False, engine='openpyxl')


hcmmap_columns_to_string = {
    'PersonalNumber HCM': str,
    'ServiceResourceName SF': str,
    'PersonalNumber SF': str,
    'PersonalNumber': str
}
hcm_map = pd.read_excel('hcm_mapping.xlsx', engine='openpyxl', dtype=hcmmap_columns_to_string)
hcm_map['PersonalNumber HCM'] = hcm_map['PersonalNumber HCM'].astype(str)
hcm_map['PersonalNumber HCM'] = hcm_map['PersonalNumber HCM'].str.strip()

hcm_file = 'HCMShifts.csv'
hcm_columns_to_string = {
    'Shop[Shop Code - Descr]': str,
    'Unique Employee[Employee Full Name]': str,
    'Unique Employee[Employee Person Number]': str
}
HCMdata = pd.read_csv(hcm_file, engine='python', dtype=hcm_columns_to_string, usecols=[
    'Shop[Shop Code - Descr]', 'Unique Employee[Employee Full Name]', 'Unique Employee[Employee Person Number]',
    'Calendar[ISO Week]', 'Calendar[ISO Year]', '[Audiologist_FTE]'
])
# Filter HCMdata between start_iso_week and end_iso_week without considering the year
HCMdata = HCMdata[
    (HCMdata['Calendar[ISO Week]'] >= start_iso_week) &
    (HCMdata['Calendar[ISO Week]'] <= end_iso_week)
]

HCMdata['ShopCode'] = HCMdata['Shop[Shop Code - Descr]'].str[:3]  # Extract the ShopCode_3char from CompositeKey
HCMdata['ShopCode_pn'] = (HCMdata['ShopCode'] + '_' +  HCMdata['Unique Employee[Employee Person Number]'].astype(str))
hcm_map = hcm_map.drop_duplicates(subset=['PersonalNumber HCM', 'PersonalNumber', 'ServiceResourceName SF'])

HCMdata = pd.merge(
    HCMdata,
    hcm_map[['PersonalNumber HCM', 'PersonalNumber', 'ServiceResourceName SF']],  # Include Region, Area, and Shop[Name]
    left_on='ShopCode_pn',
    right_on='PersonalNumber HCM',
    how='left'
)

aihoa = HCMdata[HCMdata['PersonalNumber']=='34521']
print(aihoa[[ 'PersonalNumber', '[Audiologist_FTE]', 'Calendar[ISO Week]']])

# If 'PersonalNumber' is NaN, input the value from 'Unique Employee[Employee Person Number]'
HCMdata['PersonalNumber'] = HCMdata['PersonalNumber'].fillna(HCMdata['Unique Employee[Employee Person Number]'])

# If 'ServiceResourceName SF' is NaN, input the value from 'Unique Employee[Employee Full Name]'
HCMdata['ServiceResourceName SF'] = HCMdata['ServiceResourceName SF'].fillna(HCMdata['Unique Employee[Employee Full Name]'])

# Now you can check for missing values again if needed
missing_rows_after_fill = HCMdata[HCMdata['ServiceResourceName SF'].isna() | HCMdata['PersonalNumber'].isna()]

# Print the rows that still have missing values (if any)
print(missing_rows_after_fill[['ShopCode_pn', 'PersonalNumber', 'ServiceResourceName SF']])

HCMdata = pd.merge(
    HCMdata,
    region_mapping[['CODE', 'SYM']],  # Include Region, Area, and Shop[Name]
    left_on='ShopCode',
    right_on='CODE',
    how='left'
)

HCMdata = HCMdata[HCMdata['SYM']=='Y']


missing_rows = HCMdata[HCMdata['ServiceResourceName SF'].isna()][['ShopCode_pn', 'PersonalNumber', 'ServiceResourceName SF']]
# Create the 'CompositeKey' column using .loc as well
HCMdata.loc[:, 'CompositeKey'] = (HCMdata['ShopCode'].astype(str) + '_' +
    HCMdata['PersonalNumber'].astype(str).str.strip() + '_' +
    HCMdata['Calendar[ISO Year]'].astype(str) + '_' + 
    HCMdata['Calendar[ISO Week]'].astype(str)
)
# Create the 'CompositeKey' in sfshifts_merged using .loc
sfshifts_merged.loc[:, 'CompositeKey'] = (
    sfshifts_merged['PersonalNumberKey'].astype(str).str.strip() + '_' +
    sfshifts_merged['iso_year'].astype(str) + '_' + 
    sfshifts_merged['iso_week'].astype(str)
)
# Step 2: Group and sum data
HCMdata_summed = HCMdata.groupby(
    ['CompositeKey', 'Calendar[ISO Year]', 'Calendar[ISO Week]']
).agg({
    '[Audiologist_FTE]': 'sum',
    'PersonalNumber': 'first',
    'ServiceResourceName SF' : 'first'
    }).reset_index()
HCMdata_summed.head()
# Multiply the '[Audiologist_FTE]' by 40 to get the duration
HCMdata_summed['Duración HCM'] = HCMdata_summed['[Audiologist_FTE]'] * 40

# Step 3: Process SF shifts data
shift_duration_per_week = sfshifts_merged.groupby(
    ['CompositeKey']
).agg({
    'ShiftDurationHours': 'sum',
    'GT_ServiceResource__r.Name': 'first',
}).reset_index()
shift_duration_per_week.rename(columns={'ShiftDurationHours': 'Duración SF'}, inplace=True)
shift_duration_per_week['PersonalNumber'] = shift_duration_per_week['CompositeKey'].str[4:-8]
# Now you can check for missing values again if needed
missing_rows_after_fill = shift_duration_per_week[shift_duration_per_week['GT_ServiceResource__r.Name'].isna() | shift_duration_per_week['PersonalNumber'].isna()]

# Print the rows that still have missing values (if any)
print(missing_rows_after_fill[['CompositeKey', 'PersonalNumber', 'GT_ServiceResource__r.Name']])

# Step 4: Merge both datasets (without region/area/shop data yet)
all_composite_keys = pd.merge(
    shift_duration_per_week[['CompositeKey', 'Duración SF', 'PersonalNumber','GT_ServiceResource__r.Name']], 
    HCMdata_summed[['CompositeKey', 'Duración HCM', 'PersonalNumber','ServiceResourceName SF']],
    on='CompositeKey', how='outer', suffixes=('_sf', '_hcm'), indicator=True
)
all_composite_keys.head()
# Step 5: Add region, area, and shop (DESCR) mapping data based on the merged composite keys
all_composite_keys['ShopCode_3char'] = all_composite_keys['CompositeKey'].str[:3]  # Extract the ShopCode_3char from CompositeKey
all_composite_keys['shop_pn'] = all_composite_keys['CompositeKey'].str[:-8]
all_composite_keys['iso_week'] = all_composite_keys['CompositeKey'].apply(lambda x: x.split('_')[-1])
all_composite_keys.columns
# If 'PersonalNumber' is NaN, input the value from 'Unique Employee[Employee Person Number]'
all_composite_keys['Personal Number'] = all_composite_keys['PersonalNumber_hcm'].fillna(all_composite_keys['PersonalNumber_sf'])

# If 'ServiceResourceName SF' is NaN, input the value from 'Unique Employee[Employee Full Name]'
all_composite_keys['Resource Name'] = all_composite_keys['ServiceResourceName SF'].fillna(all_composite_keys['GT_ServiceResource__r.Name'])
# Normalize 'Resource_Name' column to capitalize the first letter of each word
all_composite_keys['Resource Name'] = all_composite_keys['Resource Name'].str.title()
# Now you can check for missing values again if needed
missing_rows_after_fill = all_composite_keys[all_composite_keys['Personal Number'].isna()]
missing_rows_after_fill
missing_rl = all_composite_keys[all_composite_keys['Personal Number']=='51180']
print(missing_rl[[ 'Personal Number', 'Resource Name', 'CompositeKey']])

# Step 6: Final Calculations and Fill Missing Values
all_composite_keys['Diferencia de hcm duración'] = all_composite_keys['Duración SF'].fillna(0) - all_composite_keys['Duración HCM'].fillna(0)

all_composite_keys = pd.merge(
    all_composite_keys,
    region_mapping[['CODE', 'AREA','REGION', 'DESCR','SYM']],  # Include Region, Area, and Shop[Name]
    left_on='ShopCode_3char',
    right_on='CODE',
    how='left'
)

all_composite_keys.rename(columns={
    'CompositeKey': 'Clave compuesta',
    'ShopCode_3char': 'Shop Code',
    'CODE': 'Code',
    'REGION': 'Region',
    'AREA': 'Area',
    'DESCR':'Shop Name',
}, inplace=True)
all_composite_keys

missing_region_rows = all_composite_keys[all_composite_keys['Region'].isna()]
print(missing_region_rows)
missing_region_rows['Shop Code'].unique()
# Remove rows where REGION is blank (i.e., NaN)
all_composite_keys = all_composite_keys[all_composite_keys['Region'].notna()]
all_composite_keys.columns
# Fill NaN values in the following columns with 0
all_composite_keys[['Duración SF', 'Duración HCM', 'Diferencia de hcm duración']] = all_composite_keys[['Duración SF', 'Duración HCM', 'Diferencia de hcm duración']].fillna(0)
# Find the duplicated rows based on 'Resource_Name' and 'iso_week'
duplicates = all_composite_keys[all_composite_keys.duplicated(subset=['Clave compuesta'], keep=False)]
# Display the first few rows of the duplicates
print(duplicates.head())
all_composite_keys.drop(columns=['PersonalNumber_sf', 'PersonalNumber_hcm','GT_ServiceResource__r.Name', '_merge', 'SYM'], inplace=True)
all_composite_keys.columns
# Step 7: Save the result to Excel
output_file_path1 = 'hcm_sf_merged.xlsx'
all_composite_keys.to_excel(output_file_path1, index=False, engine='openpyxl')


def load_and_merge_files(directory, file_pattern):
    # List to store dataframes
    all_dfs = []
    
    # Regular expression to match the filename pattern flexibly
    pattern = re.compile(r"1039963987_.*_1_1_ *\.xlsx")

    # List files in directory for debugging
    print("Files in directory:", os.listdir(directory))

    # Scan for files matching the regex pattern
    for filename in os.listdir(directory):
        print(f"Checking {filename}...")
        if pattern.match(filename):
            file_path = os.path.join(directory, filename)
            print(f"Loading {filename}")
            # Load the file, skipping the first 4 rows, using the 5th row as header
            df = pd.read_excel(file_path, header=6)
            df['ID RH'] = df['ID RH'].astype(str).str.strip()
            all_dfs.append(df)
    
    # Merge all dataframes into one
    if all_dfs:
        clock = pd.concat(all_dfs, ignore_index=True)
        print("All files merged successfully.")
        return clock
    else:
        print("No files found matching the pattern.")
        return None

# Initial load of files
directory = 'files'
file_pattern = "1039963987_*_1_1_"
clock = load_and_merge_files(directory, file_pattern)


# Function to check for new files and merge them
def check_for_new_files(clock, directory, file_pattern):
    # Track loaded files
    loaded_files = set(clock['filename'].unique()) if clock is not None else set()
    
    while True:
        for filename in os.listdir(directory):
            if filename.startswith(file_pattern) and filename.endswith('.xlsx'):
                if filename not in loaded_files:
                    print(f"New file detected: {filename}")
                    file_path = os.path.join(directory, filename)
                    new_df = pd.read_excel(file_path, header=6)
                    new_df['ID RH'] = new_df['ID RH'].astype(str).str.strip()
                    # Add filename to track
                    new_df['filename'] = filename
                    # Merge new data
                    clock = pd.concat([clock, new_df], ignore_index=True)
                    loaded_files.add(filename)
                    print(f"{filename} has been merged.")
                    
        # Optional sleep time to avoid constant disk reads
        time.sleep(10)  # Check every 10 seconds


clock.columns
#Clock-in-out
#clock = '1039467394_4_1_1_ .xlsx'
# Load the Excel file, skipping the first 4 rows and using the 5th row as headers
#clock = pd.read_excel(clock, header=6)
#clock['ID RH'] = clock['ID RH'].astype(str).str.strip()
# Convert to string, handling NaN and removing any .0 from floats
clock['ID RH'] = clock['ID RH'].astype(str).replace(r'\.0$', '', regex=True).replace('nan', '')

# Assuming 'df' is the DataFrame and 'Id.Empleado' is the column to check for duplicates
duplicates = clock[clock.duplicated(subset=['ID RH', 'Fecha y hora fichaje/declarac.'], keep=False)]

# Display the duplicate rows
print(duplicates)

# Step 1: Ensure that the 'Fecha y hora fichaje' column is in datetime format
clock['Fecha y hora fichaje'] = pd.to_datetime(clock['Fecha y hora fichaje/declarac.'])

# Step 2: Sort the data by 'Id.Empleado' (ID RH) and 'Fecha y hora fichaje'
clock_sorted = clock.sort_values(by=['ID RH', 'Fecha y hora fichaje'])
clock_sorted.columns
check_data = clock_sorted[(clock_sorted['ID RH'] == '41884') & (clock_sorted['Fecha y hora fichaje'].dt.date == pd.to_datetime('2024-10-07').date())]
# Display the relevant columns for verification
check_data[['ID RH', 'Fecha y hora fichaje']]
clock_sorted['date_only'] = clock_sorted['Fecha y hora fichaje'].dt.date
# Step 3: Assign alternating "Clock In" and "Clock Out" labels within each group of 'ID RH'
clock_sorted['clock_type'] = clock_sorted.groupby(['ID RH', 'date_only']).cumcount() % 2
clock_sorted['clock_type'] = clock_sorted['clock_type'].map({0: 'Clock In', 1: 'Clock Out'})
clock_sorted.drop(columns=['date_only'], inplace=True)
# Step 4: Calculate the time difference only for "Clock In" and the next "Clock Out"
# Shift the 'Fecha y hora fichaje' column to calculate time difference for Clock In to next Clock Out
clock_sorted['next_fichaje'] = clock_sorted.groupby('ID RH')['Fecha y hora fichaje'].shift(-1)
clock_sorted['time_diff'] = clock_sorted['next_fichaje'] - clock_sorted['Fecha y hora fichaje']
# Step 5: Keep only the rows that are "Clock In" to calculate working hours
clock_in = clock_sorted[clock_sorted['clock_type'] == 'Clock In'].copy()
check_data = clock_sorted[(clock_sorted['ID RH'] == '41884') & (clock_sorted['Fecha y hora fichaje'].dt.date == pd.to_datetime('2024-10-07').date())]
# Display the relevant columns for verification
check_data[['ID RH', 'Fecha y hora fichaje']]
# Filter for the specific PersonalNumberKey and Date
check_data = clock_sorted[(clock_sorted['ID RH'] == '41884') & (clock_sorted['Fecha y hora fichaje'].dt.date == pd.to_datetime('2024-10-07').date())]
# Display the relevant columns for verification
check_data[['ID RH', 'Fecha y hora fichaje', 'clock_type', 'next_fichaje', 'time_diff']]
# Step 6: Convert time differences to hours (optional)
clock_in['hours_worked'] = clock_in.apply(
    lambda row: row['time_diff'].total_seconds() / 3600 
                if pd.notnull(row['time_diff']) and row['next_fichaje'].date() == row['Fecha y hora fichaje'].date() 
                else 'NC', 
    axis=1
) 
# Filter for the specific PersonalNumberKey and Date
check_data = clock_in[(clock_in['ID RH'] == '41884') & (clock_in['Fecha y hora fichaje'].dt.date == pd.to_datetime('2024-10-07').date())]

# Display the relevant columns for verification
check_data[['ID RH', 'Fecha y hora fichaje', 'clock_type', 'next_fichaje', 'time_diff', 'hours_worked']]

# Step 7: Apply the NC logic:
# Check for any 'Clock Out' entries for the day and employee
clock_in['Date'] = clock_in['Fecha y hora fichaje'].dt.date
has_clock_out = clock_sorted.groupby(['ID RH', clock_sorted['Fecha y hora fichaje'].dt.date])['clock_type'].apply(lambda x: (x == 'Clock Out').any()).reset_index()
has_clock_out.columns = ['ID RH', 'Date', 'has_clock_out']

# Identify if there is at least one 'Clock Out' on the same day for each ID RH
clock_out_by_day = clock_in[(clock_in['clock_type'] == 'Clock Out')].groupby(['ID RH', 'Date']).size().reset_index(name='clock_out_count')
clock_out_by_day['has_clock_out'] = clock_out_by_day['clock_out_count'] > 0
clock_in = clock_in.merge(clock_out_by_day[['ID RH', 'Date', 'has_clock_out']], on=['ID RH', 'Date'], how='left')

# Apply NC logic based on the presence of a valid clock-out for the day
clock_in['is_nc'] = clock_in.apply(lambda row: 1 if row['hours_worked'] == 'NC' and not row['has_clock_out'] else 0, axis=1)
clock_in['hours_worked'] = clock_in.apply(
    lambda row: 0 if row['is_nc'] == 1 else row['hours_worked'], 
    axis=1
)

# Replace 'NC' with 0 and convert to float explicitly
clock_in['hours_worked_numeric'] = clock_in['hours_worked'].replace('NC', 0)
clock_in['hours_worked_numeric'] = clock_in['hours_worked_numeric'].astype(float)

# Drop the 'has_clock_out' column if it’s no longer needed
clock_in.drop(columns='has_clock_out', inplace=True)
clock_in
# Filter for the specific PersonalNumberKey and Date
check_data = clock_in[(clock_in['ID RH'] == '41884') & (clock_in['Fecha y hora fichaje'].dt.date == pd.to_datetime('2024-10-07').date())]

# Display the relevant columns for verification
check_data[['ID RH', 'Fecha y hora fichaje', 'clock_type', 'next_fichaje', 'time_diff', 'hours_worked', 'hours_worked_numeric']]


clock_in[['hours_worked_numeric', 'hours_worked']].head() 
clock_in['Shop Name'] = clock_in['Nombre unidad org.'].str.replace('ES - SHOP - ', '', regex=False)
clock_in['Shop Name'] = clock_in['Shop Name'].str.strip()
clock_in['Shop Name'] = clock_in['Shop Name'].replace("L’HOSPITALET DE LLOBREGAT - JUST OLIVERES", "L'HOSPITALET DE LLOBREGAT - JUST OLIVERES")

clock_in['is_nc'] = clock_in['hours_worked'].apply(lambda x: 1 if x == 'NC' else 0)
# Step 8: Group by 'ID RH' and 'Fecha y hora fichaje' date to adjust 'is_nc' based on any valid clock-out during the same day
clock_in['date_only'] = clock_in['Fecha y hora fichaje'].dt.date
clock_in = clock_in.merge(
    clock_in.groupby(['ID RH', 'date_only'], as_index=False).agg(
        has_valid_pair=('is_nc', lambda x: 0 if 0 in x.values else 1)
    ),
    on=['ID RH', 'date_only']
)
# Update 'is_nc' to be 0 if there's a valid clock-out pair on that day
clock_in['is_nc'] = clock_in.apply(lambda row: 0 if row['has_valid_pair'] == 0 else row['is_nc'], axis=1)

nc_rows_filtered = clock_in[clock_in['is_nc'] == 1][['is_nc', 'Fecha y hora fichaje', 'clock_type', 'next_fichaje']]

nc_rows_filtered[['is_nc', 'Fecha y hora fichaje', 'clock_type','next_fichaje' ]].head(20)
clock_in.drop(columns=['date_only', 'has_valid_pair'], inplace=True)

total_hours_per_employee = pd.merge(
    clock_in,
    region_mapping[['CODE', 'REGION', 'AREA', 'DESCR','SYM']],  # Include Region, Area, and Shop[Name]
    left_on='Shop Name',
    right_on='DESCR',
    how='left'
)
print(total_hours_per_employee[total_hours_per_employee['Shop Name'] == "L'HOSPITALET DE LLOBREGAT - JUST OLIVERES"])
total_hours_per_employee=total_hours_per_employee[total_hours_per_employee['SYM']=='Y']
total_hours_per_employee['ISO Year'] = total_hours_per_employee['Fecha y hora fichaje'].dt.isocalendar().year
total_hours_per_employee['ISO Week'] = total_hours_per_employee['Fecha y hora fichaje'].dt.isocalendar().week
total_hours_per_employee['Fecha y hora fichaje'] = pd.to_datetime(total_hours_per_employee['Fecha y hora fichaje'])
total_hours_per_employee['Date'] = total_hours_per_employee['Fecha y hora fichaje'].dt.date
# Group by with the new column and sum it, while keeping the original 'hours_worked' column
total_hours_per_employee_daily = total_hours_per_employee.groupby(
    ['Date', 'ID RH'], as_index=False
).agg({
    'hours_worked_numeric': 'sum',  # Summing the hours worked, with 'NC' as 0
    'Shop Name': 'first',   
    'CODE': 'first',        
    'REGION': 'first',
    'AREA': 'first',
    'DESCR': 'first',
    'SYM': 'first',
    'is_nc': 'max'  # Check if any entry within the group was 'NC'

})
print(total_hours_per_employee_daily.head())
# Filter for rows where 'hours_worked' is 'NC'
total_hours_per_employee_daily['hours_worked'] = total_hours_per_employee_daily.apply(
    lambda row: 'NC' if row['is_nc'] == 1 else row['hours_worked_numeric'], axis=1
)
# Drop the temporary 'is_nc' column as it is no longer needed
total_hours_per_employee_daily.drop(columns='is_nc', inplace=True)
total_hours_per_employee_daily[['hours_worked_numeric', 'hours_worked']].head()
total_hours_per_employee_daily.columns
# Filter for the specific Resource number and Date
check_data = total_hours_per_employee_daily[(total_hours_per_employee_daily['ID RH'] == '41884') & (total_hours_per_employee_daily['Date'] == pd.to_datetime('2024-10-07').date())]
# Display the relevant columns for verification
check_data[['Date', 'hours_worked', 'hours_worked_numeric']]

total_hours_per_employee_daily['Date'] = pd.to_datetime(total_hours_per_employee_daily['Date']).dt.date
sfshifts_merged['ShiftDate'] = pd.to_datetime(sfshifts_merged['ShiftDate']).dt.date
sfshifts_merged[sfshifts_merged['PersonalNumberKey'] == '003_11126'].head()
sfshifts_merged['PersonalNumber'] = sfshifts_merged['PersonalNumberKey'].str[4:]
sfshifts_merged_per_emp = sfshifts_merged.groupby(['PersonalNumber', 'ShiftDate'])[[ 'ShiftDurationHours', 'AbsenceDurationHours', 'ShiftDurationHoursAdjusted']].sum().reset_index()
sfshifts_merged_per_emp[sfshifts_merged_per_emp['PersonalNumber']=='31992']

# Step 4: Merge both datasets (without region/area/shop data yet)
clockin_merged = pd.merge(
    total_hours_per_employee_daily[['ID RH', 'Date', 'hours_worked', 'hours_worked_numeric']], 
    sfshifts_merged_per_emp[['PersonalNumber', 'ShiftDate', 'ShiftDurationHours', 'AbsenceDurationHours', 'ShiftDurationHoursAdjusted']],
    left_on=['ID RH', 'Date'],    
    right_on=['PersonalNumber', 'ShiftDate'], 
    how='outer', 
    suffixes=('_sf', '_act'), 
    indicator=True 
)
clockin_merged
hcm_map_active = hcm_map[hcm_map['Active'] == True].copy()
hcm_map_active
clockin_merged.head()
clockin_merged = pd.merge(
    clockin_merged,
    hcm_map_active[['PersonalNumber', 'ServiceResourceName SF','PersonalNumber SF']],
    left_on='PersonalNumber',
    right_on='PersonalNumber',
    how='left'
)
clockin_merged['ShopCode'] = clockin_merged['PersonalNumber SF'].str[:3]
clockin_merged = pd.merge(
    clockin_merged,
    region_mapping[['CODE', 'AREA','REGION', 'DESCR','SYM']],  # Include Region, Area, and Shop[Name]
    left_on='ShopCode',
    right_on='CODE',
    how='left'
)
clockin_merged = clockin_merged[clockin_merged['SYM']=='Y']
clockin_merged.columns
clockin_merged[[ 'hours_worked','ShiftDurationHours', 'AbsenceDurationHours', 'ShiftDurationHoursAdjusted']] = clockin_merged[['hours_worked','ShiftDurationHours', 'AbsenceDurationHours', 'ShiftDurationHoursAdjusted']].fillna(0)
# Now you can check for missing values again if needed
clockin_merged['Resource Name'] = clockin_merged['ServiceResourceName SF']
clockin_merged['Resource Name'] = clockin_merged['Resource Name'].str.title()
clockin_merged['Date'] = clockin_merged['Date'].fillna(clockin_merged['ShiftDate'])
clockin_merged['Date'] = pd.to_datetime(clockin_merged['Date'])
clockin_merged['weekday'] = clockin_merged['Date'].dt.day_name()
clockin_merged['iso_week'] = clockin_merged['Date'].dt.isocalendar().week
missing_rows_after_fill = clockin_merged[clockin_merged['Resource Name'].isna()]

clockin_merged.columns
clockin_merged.rename(columns={
    'ShopCode': 'Shop Code',
    'CODE': 'Code',
    'REGION': 'Region',
    'AREA': 'Area',
    'DESCR':'Shop[Name]',
}, inplace=True)
clockin_merged.columns
clockin_merged.drop(columns=['ServiceResourceName SF', '_merge', 'SYM', 'ID RH'], inplace=True)
clockin_merged['Diferencia de act duración'] = clockin_merged['ShiftDurationHoursAdjusted'].fillna(0) - clockin_merged['hours_worked_numeric'].fillna(0)
clockin_merged[['hours_worked_numeric', 'hours_worked']].head() 
clockin_merged = clockin_merged.drop_duplicates(subset=['PersonalNumber', 'Date'])
duplicates = clockin_merged[clockin_merged.duplicated(subset=['PersonalNumber', 'Date'])]
duplicates
output_file_path3 = 'clock.xlsx'
clockin_merged.to_excel(output_file_path3, index=False, engine='openpyxl')
